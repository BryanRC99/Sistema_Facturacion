from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter    
from django.shortcuts import render, redirect, get_object_or_404
from .models import Proveedor
from django.template.loader import render_to_string
from weasyprint import HTML
from django.utils import timezone
from django.contrib import messages
from usuarios.decorators import group_required



@group_required(["Vendedor", "SuperAdmin"])
def lista_proveedores(request):
    proveedores = Proveedor.objects.all()

    # Para mostrar el texto del choice (RUC, Cédula, etc.)
    for p in proveedores:
        p.tipo_identificacion_display = p.get_tipo_identificacion_display()

    return render(request, 'proveedores/lista.html', {
        'proveedores': proveedores
    })

@group_required(["Vendedor", "SuperAdmin"])
def crear_proveedor(request):
    if request.method == 'POST':
        form_data = request.POST

        tipo_identificacion = request.POST.get('tipo_identificacion', '').strip()
        identificacion = request.POST.get('identificacion', '').strip()
        nombre_razon_social = request.POST.get('nombre_razon_social', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        celular = request.POST.get('celular', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        correo = request.POST.get('correo', '').strip()

        ciudad = (request.POST.get('ciudad') or '').strip() or None
        contacto_nombre = (request.POST.get('contacto_nombre') or '').strip() or None
        contacto_telefono = (request.POST.get('contacto_telefono') or '').strip() or None
        observaciones = (request.POST.get('observaciones') or '').strip() or None

        # =========================
        # VALIDACIONES
        # =========================
        if not tipo_identificacion:
            messages.error(request, "Debes seleccionar el tipo de identificación.")
            return render(request, 'proveedores/crear.html', {"form_data": form_data})

        if not identificacion:
            messages.error(request, "La identificación es obligatoria.")
            return render(request, 'proveedores/crear.html', {"form_data": form_data})

        if not nombre_razon_social:
            messages.error(request, "El nombre o razón social es obligatorio.")
            return render(request, 'proveedores/crear.html', {"form_data": form_data})

        if not direccion:
            messages.error(request, "La dirección es obligatoria.")
            return render(request, 'proveedores/crear.html', {"form_data": form_data})

        # Duplicado de identificación (ignora mayúsculas/minúsculas)
        if Proveedor.objects.filter(identificacion__iexact=identificacion).exists():
            messages.warning(request, "Ya existe un proveedor con esta identificación.")
            return render(request, 'proveedores/crear.html', {"form_data": form_data})

        # Correo básico (si se envía)
        if correo and ("@" not in correo or "." not in correo):
            messages.error(request, "El correo no parece válido.")
            return render(request, 'proveedores/crear.html', {"form_data": form_data})

        # =========================
        # CREAR
        # =========================
        Proveedor.objects.create(
            tipo_identificacion=tipo_identificacion,
            identificacion=identificacion,
            nombre_razon_social=nombre_razon_social,
            telefono=telefono,
            celular=celular,
            direccion=direccion,
            correo=correo,
            ciudad=ciudad,
            contacto_nombre=contacto_nombre,
            contacto_telefono=contacto_telefono,
            observaciones=observaciones,
        )

        messages.success(request, "Proveedor creado correctamente.")
        return redirect('lista_proveedores')

    return render(request, 'proveedores/crear.html')

@group_required(["Vendedor", "SuperAdmin"])
def editar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)

    if request.method == 'POST':
        form_data = request.POST

        tipo_identificacion = request.POST.get('tipo_identificacion', '').strip()
        nueva_identificacion = request.POST.get('identificacion', '').strip()
        nombre_razon_social = request.POST.get('nombre_razon_social', '').strip()
        telefono = (request.POST.get('telefono') or '').strip()
        celular = (request.POST.get('celular') or '').strip()
        direccion = request.POST.get('direccion', '').strip()
        correo = (request.POST.get('correo') or '').strip()

        ciudad = (request.POST.get('ciudad') or '').strip() or None
        # CONSISTENTE con crear_proveedor:
        contacto_nombre = (request.POST.get('contacto') or '').strip() or None
        contacto_telefono = (request.POST.get('contacto_telefono') or '').strip() or None
        observaciones = (request.POST.get('observaciones') or '').strip() or None

        # =========================
        # VALIDACIONES
        # =========================
        if not tipo_identificacion:
            messages.error(request, "Debes seleccionar el tipo de identificación.")
            return render(request, 'proveedores/editar.html', {
                "proveedor": proveedor,
                "form_data": form_data
            })

        if not nueva_identificacion:
            messages.error(request, "La identificación es obligatoria.")
            return render(request, 'proveedores/editar.html', {
                "proveedor": proveedor,
                "form_data": form_data
            })

        if not nombre_razon_social:
            messages.error(request, "El nombre o razón social es obligatorio.")
            return render(request, 'proveedores/editar.html', {
                "proveedor": proveedor,
                "form_data": form_data
            })

        if not direccion:
            messages.error(request, "La dirección es obligatoria.")
            return render(request, 'proveedores/editar.html', {
                "proveedor": proveedor,
                "form_data": form_data
            })

        # Evitar duplicado de identificación
        if Proveedor.objects.filter(identificacion__iexact=nueva_identificacion).exclude(id=proveedor.id).exists():
            messages.warning(request, "Otro proveedor ya está utilizando esta identificación.")
            return render(request, 'proveedores/editar.html', {
                "proveedor": proveedor,
                "form_data": form_data
            })

        if correo and ("@" not in correo or "." not in correo):
            messages.error(request, "El correo no parece válido.")
            return render(request, 'proveedores/editar.html', {
                "proveedor": proveedor,
                "form_data": form_data
            })

        # =========================
        # GUARDAR
        # =========================
        proveedor.tipo_identificacion = tipo_identificacion
        proveedor.identificacion = nueva_identificacion
        proveedor.nombre_razon_social = nombre_razon_social
        proveedor.telefono = telefono
        proveedor.celular = celular
        proveedor.direccion = direccion
        proveedor.correo = correo
        proveedor.ciudad = ciudad
        proveedor.contacto_nombre = contacto_nombre
        proveedor.contacto_telefono = contacto_telefono
        proveedor.observaciones = observaciones
        proveedor.save()

        messages.success(request, "Proveedor actualizado correctamente.")
        return redirect('lista_proveedores')

    return render(request, 'proveedores/editar.html', {
        'proveedor': proveedor
    })

@group_required(["Vendedor", "SuperAdmin"])
def eliminar_proveedor(request, id):
    proveedor = get_object_or_404(Proveedor, id=id)
    proveedor.delete()
    messages.success(request, "Proveedor eliminado correctamente.")
    return redirect('lista_proveedores')


@group_required(["Vendedor", "SuperAdmin"])
def export_proveedores_pdf(request):
    proveedores = Proveedor.objects.all().order_by("id")

    html_string = render_to_string(
        "proveedores/export_pdf.html",
        {
            "proveedores": proveedores,
            "fecha": timezone.localtime().strftime("%d/%m/%Y %H:%M"),
        }
    )

    pdf_bytes = HTML(
        string=html_string,
        base_url=request.build_absolute_uri("/")
    ).write_pdf()

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="proveedores.pdf"'
    return response


@group_required(["Vendedor", "SuperAdmin"])
def export_proveedores_excel(request):
    proveedores = Proveedor.objects.all().order_by("id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    headers = [
        "#",
        "Tipo Identificación",
        "Identificación",
        "Nombre / Razón Social",
        "Teléfono",
        "Celular",
        "Ciudad",
        "Contacto Nombre",
        "Contacto Teléfono",
        "Dirección",
        "Correo",
        "Observaciones",
    ]
    ws.append(headers)

    # Header style
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for i, p in enumerate(proveedores, start=1):
        ws.append([
            i,
            p.tipo_identificacion,
            p.identificacion,
            p.nombre_razon_social,
            p.telefono,
            p.celular,
            p.ciudad or "",
            p.contacto_nombre or "",
            p.contacto_telefono or "",
            p.direccion,
            p.correo,
            p.observaciones or "",
        ])

    # Auto width
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)

    filename = f"proveedores_{timezone.localtime().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response