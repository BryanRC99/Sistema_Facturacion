from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from weasyprint import HTML
from django.utils import timezone  
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Categoria
from django.contrib import messages


def inicio(request):
    return HttpResponse("Bienvenido al sistema de facturación")

@login_required(login_url='/accounts/login/')
def lista_categorias(request):
    categorias = Categoria.objects.all()
    return render(request, 'categorias/lista.html', {'categorias': categorias})

def crear_categoria(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()

        if not nombre:
            messages.error(request, "El nombre de la categoría es obligatorio.")
            return redirect('crear_categoria')

        if Categoria.objects.filter(nombre__iexact=nombre).exists():
            messages.warning(request, "Ya existe una categoría con ese nombre.")
            return redirect('crear_categoria')

        Categoria.objects.create(nombre=nombre)
        messages.success(request, "Categoría creada correctamente.")
        return redirect('lista_categorias')

    return render(request, 'categorias/crear.html')

def editar_categoria(request, id):
    categoria = get_object_or_404(Categoria, id=id)

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()

        if not nombre:
            messages.error(request, "El nombre de la categoría es obligatorio.")
            return redirect('editar_categoria', id=id)

        if Categoria.objects.filter(nombre__iexact=nombre).exclude(id=id).exists():
            messages.warning(request, "Ya existe otra categoría con ese nombre.")
            return redirect('editar_categoria', id=id)

        categoria.nombre = nombre
        categoria.save()

        messages.success(request, "Categoría actualizada correctamente.")
        return redirect('lista_categorias')

    return render(request, 'categorias/editar.html', {
        'categoria': categoria
    })

def eliminar_categoria(request, id):
    categoria = get_object_or_404(Categoria, id=id)
    categoria.delete()

    messages.success(request, "Categoría eliminada correctamente.")
    return redirect('lista_categorias')

def export_categorias_pdf(request):
    categorias = Categoria.objects.all().order_by("id")

    html_string = render_to_string(
        "categorias/export_pdf.html",
        {
            "categorias": categorias,
            "fecha": timezone.localtime().strftime("%d/%m/%Y %H:%M"),
        }
    )

    pdf = HTML(
        string=html_string,
        base_url=request.build_absolute_uri("/")
    ).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="categorias.pdf"'
    return response


def export_categorias_excel(request):
    categorias = Categoria.objects.all().order_by("id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías"

    headers = ["#", "Nombre"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for i, c in enumerate(categorias, start=1):
        ws.append([i, c.nombre])

    # ancho columnas
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[col_letter])
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    filename = f"categorias_{timezone.localtime().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
