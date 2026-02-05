import os
from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from weasyprint import HTML
from django.utils import timezone    
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
import cloudinary.uploader
from django.core.exceptions import ValidationError
from .models import Producto, Categoria
from proveedores.models import Proveedor
from django.contrib import messages
from usuarios.decorators import group_required



@group_required(["Vendedor", "SuperAdmin"])
def verificar_codigo(request):
    codigo = request.GET.get("codigo", "").strip().lower()
    existe = Producto.objects.filter(codigo__iexact=codigo).exists()
    return JsonResponse({"existe": existe})

@group_required(["Vendedor", "SuperAdmin"])
def lista_productos(request):
    productos = Producto.objects.all().select_related('categoria')
    return render(request, 'productos/lista.html', {'productos': productos})

@group_required(["Vendedor", "SuperAdmin"])
def crear_producto(request):
    categorias = Categoria.objects.all()
    proveedores = Proveedor.objects.filter(estado=True)
    productos = Producto.objects.all()

    error = None
    form_data = {}

    if request.method == 'POST':
        form_data = {
            'codigo': request.POST.get('codigo', '').strip(),
            'nombre': request.POST.get('nombre', '').strip(),
            'categoria': request.POST.get('categoria'),
            'proveedor': request.POST.get('proveedor'),
            'precio': request.POST.get('precio'),
            'costo': request.POST.get('costo'),
            'stock': request.POST.get('stock'),
            'iva': request.POST.get('iva') == 'on',
            'estado': request.POST.get('estado') == 'on',
            'descripcion': request.POST.get('descripcion', '').strip(),
        }

        imagen = request.FILES.get('imagen')

        if Producto.objects.filter(codigo=form_data['codigo']).exists():
            error = "El código ingresado ya existe. Por favor ingrese uno diferente."

        elif imagen:
            formatos_permitidos = ('image/jpeg', 'image/png', 'image/webp')
            max_size = 2 * 1024 * 1024

            if imagen.content_type not in formatos_permitidos:
                error = "Formato de imagen no permitido. Use JPG, PNG o WEBP."
            elif imagen.size > max_size:
                error = "La imagen no debe superar los 2MB."

        if not error:
            Producto.objects.create(
                codigo=form_data['codigo'],
                nombre=form_data['nombre'],
                categoria_id=form_data['categoria'],
                proveedor_id=form_data['proveedor'] or None,
                precio=float(form_data['precio']),
                costo=float(form_data['costo']),
                stock=int(form_data['stock']),
                iva=form_data['iva'],
                descripcion=form_data['descripcion'],
                estado=form_data['estado'],
                imagen=imagen if imagen else None
            )
            messages.success(request, "Producto guardado correctamente.")
            return redirect('lista_productos')

        # Si hubo error (validación), lo mostramos como toast también
        messages.error(request, f"❌ {error}")

    return render(request, 'productos/crear.html', {
        'categorias': categorias,
        'proveedores': proveedores,
        'productos': productos,
        'error': error,          # (puedes dejarlo por si lo usas en el template)
        'form_data': form_data,
    })


@group_required(["Vendedor", "SuperAdmin"])
def editar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)
    categorias = Categoria.objects.all()
    proveedores = Proveedor.objects.all()

    if request.method == 'POST':
        producto.codigo = request.POST.get('codigo', '').strip()
        producto.nombre = request.POST.get('nombre', '').strip()
        producto.categoria_id = request.POST.get('categoria')
        producto.proveedor_id = request.POST.get('proveedor') or None
        producto.precio = float(request.POST.get('precio', 0) or 0)
        producto.costo = float(request.POST.get('costo', 0) or 0)
        producto.stock = int(request.POST.get('stock', 0) or 0)
        producto.iva = request.POST.get('iva') == 'on'
        producto.descripcion = request.POST.get('descripcion', '').strip()
        producto.estado = request.POST.get('estado') == 'on'

        # ===============================
        # MANEJO DE IMAGEN CLOUDINARY
        # ===============================
        imagen_nueva = request.FILES.get('imagen')

        if imagen_nueva:
            # 🔥 BORRAR IMAGEN ANTERIOR DE CLOUDINARY
            if producto.imagen:
                try:
                    cloudinary.uploader.destroy(producto.imagen.public_id)
                except Exception as e:
                    print(f"Error borrando imagen Cloudinary: {e}")
                    messages.warning(
                        request,
                        "⚠️ Se actualizó el producto, pero no se pudo borrar la imagen anterior."
                    )

            # Asignar nueva imagen
            producto.imagen = imagen_nueva

        producto.save()
        messages.success(request, "Producto actualizado correctamente.")
        return redirect('lista_productos')

    return render(request, 'productos/editar.html', {
        'producto': producto,
        'categorias': categorias,
        'proveedores': proveedores,
    })


@group_required(["Vendedor", "SuperAdmin"])
def eliminar_producto(request, id):
    producto = get_object_or_404(Producto, id=id)

    if producto.imagen:
        try:
            cloudinary.uploader.destroy(producto.imagen.public_id)
        except Exception as e:
            print(f"Error borrando imagen Cloudinary: {e}")
            messages.warning(request, "⚠️ Producto eliminado, pero no se pudo borrar la imagen en Cloudinary.")

    producto.delete()
    messages.success(request, "Producto eliminado correctamente.")
    return redirect('lista_productos')



@group_required(["Vendedor", "SuperAdmin"])
def export_productos_pdf(request):
    productos = (
        Producto.objects
        .select_related("categoria", "proveedor")
        .all()
        .order_by("id")
    )

    html_string = render_to_string(
        "productos/export_pdf.html",
        {
            "productos": productos,
            "fecha": timezone.localtime().strftime("%d/%m/%Y %H:%M"),
        }
    )

    pdf = HTML(
        string=html_string,
        base_url=request.build_absolute_uri("/")
    ).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="productos.pdf"'
    return response


@group_required(["Vendedor", "SuperAdmin"])
def export_productos_excel(request):
    productos = (
        Producto.objects
        .select_related("categoria", "proveedor")
        .all()
        .order_by("id")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = [
        "Código", "Nombre", "Categoría", "Proveedor",
        "Precio", "Costo", "Stock", "IVA", "Estado", "Descripción"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for p in productos:
        ws.append([
            p.codigo,
            p.nombre,
            p.categoria.nombre,
            p.proveedor.nombre_razon_social if p.proveedor else "",
            float(p.precio),
            float(p.costo),
            p.stock,
            "Sí" if p.iva else "No",
            "Activo" if p.estado else "Inactivo",
            p.descripcion or "",
        ])

    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        max_len = max(len(str(c.value)) if c.value else 0 for c in ws[col_letter])
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    filename = f"productos_{timezone.localtime().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
