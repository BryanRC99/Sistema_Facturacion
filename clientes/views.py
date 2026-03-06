from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from .models import Cliente
from facturacion_app.models import Factura
from django.templatetags.static import static   
from django.template.loader import render_to_string
from weasyprint import HTML
from django.utils import timezone
from usuarios.decorators import group_required
from django.contrib.auth.views import LoginView
from django.urls import reverse_lazy
from django.urls import reverse
from .decorators import cliente_required
from django.contrib.auth import get_user_model
from django.contrib.auth import update_session_auth_hash
from django.db import transaction
from django.contrib.auth.models import Group


from clientes.models import Cliente, ClienteAccount
from usuarios.models import PerfilUsuario, Rol
from .forms import ClienteCambioPasswordForm, ClienteDatosForm
from facturacion_app.models import Factura, DetalleFactura


#group_required(["Vendedor", "SuperAdmin"])
def lista_clientes(request):
    print("ENTRÓ A LISTA CLIENTES")
    clientes = Cliente.objects.all()
    for c in clientes:
        c.tipo_identificacion_display = c.get_tipo_identificacion_display()
    return render(request, 'clientes/lista.html', {'clientes': clientes})

@group_required(["Vendedor", "SuperAdmin"])
def crear_cliente(request):
    if request.method == "POST":
        form_data = request.POST

        tipo_identificacion = request.POST.get("tipo_identificacion", "").strip()
        identificacion = request.POST.get("identificacion", "").strip()
        nombre_razon_social = request.POST.get("nombre_razon_social", "").strip()
        telefono = request.POST.get("telefono", "").strip()
        celular = request.POST.get("celular", "").strip()
        direccion = request.POST.get("direccion", "").strip()
        correo = request.POST.get("correo", "").strip().lower()

        # =========================
        # VALIDACIONES
        # =========================
        if not tipo_identificacion:
            messages.error(request, "Debes seleccionar el tipo de identificación.")
            return render(request, "clientes/crear.html", {"form_data": form_data})

        if not identificacion:
            messages.error(request, "La identificación es obligatoria.")
            return render(request, "clientes/crear.html", {"form_data": form_data})

        if not nombre_razon_social:
            messages.error(request, "El nombre o razón social es obligatorio.")
            return render(request, "clientes/crear.html", {"form_data": form_data})

        # Cliente duplicado
        if Cliente.objects.filter(identificacion__iexact=identificacion).exists():
            messages.warning(request, "Ya existe un cliente con esta identificación.")
            return render(request, "clientes/crear.html", {"form_data": form_data})

        # Validación simple correo
        if correo and ("@" not in correo or "." not in correo):
            messages.error(request, "El correo no parece válido.")
            return render(request, "clientes/crear.html", {"form_data": form_data})

        # =========================
        # CREAR CLIENTE + CUENTA
        # =========================
        User = get_user_model()
        username = identificacion  # tal como quieres
        password_inicial = identificacion

        # Si YA existe un usuario con esa identificación, evitamos conflicto
        # (podría ser un usuario interno con ese username)
        if User.objects.filter(username=username).exists():
            messages.error(
                request,
                "Ya existe un usuario con ese nombre (identificación). "
                "No se puede crear la cuenta del cliente con esa identificación como usuario."
            )
            return render(request, "clientes/crear.html", {"form_data": form_data})

        try:
            with transaction.atomic():
                # 1) Crear cliente
                cliente = Cliente.objects.create(
                    tipo_identificacion=tipo_identificacion,
                    identificacion=identificacion,
                    nombre_razon_social=nombre_razon_social,
                    telefono=telefono or None,
                    celular=celular or None,
                    direccion=direccion or None,
                    correo=correo or None,
                    estado=True,
                )

                # 2) Crear user del cliente
                user = User.objects.create_user(
                    username=username,
                    password=password_inicial,
                    email=correo or "",
                )

                from django.contrib.auth.models import Group
                grupo_clientes = Group.objects.get(name="Clientes")
                user.groups.add(grupo_clientes)

                # 3) Asegurar rol CLIENTE
                rol_cliente, _ = Rol.objects.get_or_create(nombre="CLIENTE")
                PerfilUsuario.objects.get_or_create(usuario=user, defaults={"rol": rol_cliente})

                # 4) Vincular ClienteAccount y marcar cambio obligatorio
                ClienteAccount.objects.create(
                    cliente=cliente,
                    user=user,
                    activo=True,
                    debe_cambiar_password=True,
                )

        except Exception as e:
            messages.error(request, f"Ocurrió un error al crear el cliente: {e}")
            return render(request, "clientes/crear.html", {"form_data": form_data})

        messages.success(
            request,
            "Cliente creado correctamente. Se creó su acceso al portal: "
            "usuario = identificación, contraseña = identificación (debe cambiarla al entrar)."
        )
        return redirect("lista_clientes")

    return render(request, "clientes/crear.html")


@group_required(["Vendedor", "SuperAdmin"])
def editar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)

    if request.method == 'POST':
        form_data = request.POST

        tipo_identificacion = request.POST.get('tipo_identificacion', '').strip()
        nueva_identificacion = request.POST.get('identificacion', '').strip()
        nombre_razon_social = request.POST.get('nombre_razon_social', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        celular = request.POST.get('celular', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        correo = request.POST.get('correo', '').strip()

        # =========================
        # VALIDACIONES
        # =========================
        if not tipo_identificacion:
            messages.error(request, "Debes seleccionar el tipo de identificación.")
            return render(request, 'clientes/editar.html', {
                "cliente": cliente,
                "form_data": form_data
            })

        if not nueva_identificacion:
            messages.error(request, "La identificación es obligatoria.")
            return render(request, 'clientes/editar.html', {
                "cliente": cliente,
                "form_data": form_data
            })

        if not nombre_razon_social:
            messages.error(request, "El nombre o razón social es obligatorio.")
            return render(request, 'clientes/editar.html', {
                "cliente": cliente,
                "form_data": form_data
            })

        # Que no exista otro con la misma identificación
        if Cliente.objects.filter(identificacion__iexact=nueva_identificacion).exclude(id=cliente.id).exists():
            messages.warning(request, "Otro cliente ya está utilizando esta identificación.")
            return render(request, 'clientes/editar.html', {
                "cliente": cliente,
                "form_data": form_data
            })

        if correo and ("@" not in correo or "." not in correo):
            messages.error(request, "El correo no parece válido.")
            return render(request, 'clientes/editar.html', {
                "cliente": cliente,
                "form_data": form_data
            })

        # =========================
        # GUARDAR
        # =========================
        cliente.tipo_identificacion = tipo_identificacion
        cliente.identificacion = nueva_identificacion
        cliente.nombre_razon_social = nombre_razon_social
        cliente.telefono = telefono
        cliente.celular = celular
        cliente.direccion = direccion
        cliente.correo = correo

        # Auditoría (lo mantengo tal cual lo tenías)
        cliente._auditoria_user = request.user

        cliente.save()
        messages.success(request, "Cliente actualizado correctamente.")
        return redirect('lista_clientes')

    return render(request, 'clientes/editar.html', {'cliente': cliente})



@group_required(["Vendedor", "SuperAdmin"])
def eliminar_cliente(request, id):
    cliente = get_object_or_404(Cliente, id=id)
    cliente.delete()
    messages.success(request, "Cliente eliminado correctamente.")
    return redirect('lista_clientes')



@group_required(["Vendedor", "SuperAdmin"])
def export_clientes_pdf(request):
    clientes = Cliente.objects.all().order_by("id")

    html_string = render_to_string(
        "clientes/export_pdf.html",  # ✅ coincide con tu estructura
        {
            "clientes": clientes,
            "fecha": timezone.localtime().strftime("%d/%m/%Y %H:%M"),
        }
    )

    pdf_bytes = HTML(
        string=html_string,
        base_url=request.build_absolute_uri("/") 
    ).write_pdf()

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="clientes.pdf"'
    return response


@group_required(["Vendedor", "SuperAdmin"])
def export_clientes_excel(request):
    clientes = Cliente.objects.all().order_by("id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Encabezados (según tus campos reales)
    headers = [
        "#",
        "Tipo Identificación",
        "Identificación",
        "Nombre / Razón Social",
        "Teléfono",
        "Celular",
        "Dirección",
        "Correo",
    ]
    ws.append(headers)

    # Estilo header
    header_fill = PatternFill("solid", fgColor="1F2937")  # gris oscuro
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Congelar encabezado y filtro
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Filas
    for i, c in enumerate(clientes, start=1):
        ws.append([
            i,
            c.tipo_identificacion,
            c.identificacion,
            c.nombre_razon_social,
            c.telefono,
            c.celular,
            c.direccion,
            c.correo,
        ])

    # Ajustar ancho de columnas
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Nombre archivo con fecha
    filename = f"clientes_{timezone.localtime().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # Respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

class ClienteLoginView(LoginView):
    template_name = "clientes/login_cliente.html"

    def form_valid(self, form):
        user = form.get_user()

        # 🔒 Solo permitir grupo Clientes
        if not user.groups.filter(name="Clientes").exists():
            messages.error(
                self.request,
                "Este acceso es exclusivo para clientes."
            )
            return redirect("login_cliente")

        return super().form_valid(form)

    # 🔥 Forzar redirección correcta
    def get_success_url(self):
        return reverse_lazy("portal_cliente")
    

#@cliente_required
def portal(request):
    cliente = request.user.cliente_account.cliente

    total_facturas = Factura.objects.filter(cliente=cliente).count()
    ultimas = Factura.objects.filter(cliente=cliente).order_by("-fecha")[:5]

    #if request.user.cliente_account.debe_cambiar_password:
        #return redirect("cambiar_password_cliente")

    return render(request, "clientes/portal.html", {
        "cliente": cliente,
        "total_facturas": total_facturas,
        "ultimas": ultimas,
    })


@cliente_required
def cambiar_password(request):

    if request.method == "POST":
        form = ClienteCambioPasswordForm(request.POST)

        if form.is_valid():
            actual = form.cleaned_data["password_actual"]
            nueva = form.cleaned_data["nueva_password"]

            if not request.user.check_password(actual):
                messages.error(request, "La contraseña actual es incorrecta.")
            else:
                request.user.set_password(nueva)
                request.user.save()

                # Mantener sesión iniciada
                update_session_auth_hash(request, request.user)

                # Marcar que ya cambió si usas ese campo
                acc = request.user.cliente_account
                if acc.debe_cambiar_password:
                    acc.debe_cambiar_password = False
                    acc.save(update_fields=["debe_cambiar_password"])

                messages.success(request, "Contraseña actualizada correctamente.")
                return redirect("portal_cliente")
    else:
        form = ClienteCambioPasswordForm()

    return render(request, "clientes/cambiar_password.html", {
        "form": form
    })


@cliente_required
def mis_facturas(request):
    cliente = request.user.cliente_account.cliente

    q = (request.GET.get("q") or "").strip()
    estado = (request.GET.get("estado") or "").strip()

    facturas = Factura.objects.filter(cliente=cliente).order_by("-fecha")

    if q:
        facturas = facturas.filter(numero__icontains=q)

    if estado:
        facturas = facturas.filter(estado__iexact=estado)

    return render(request, "clientes/mis_facturas.html", {
        "cliente": cliente,
        "facturas": facturas,
        "q": q,
        "estado": estado,
    })


@cliente_required
def factura_detalle_cliente(request, factura_id):
    cliente = request.user.cliente_account.cliente

    factura = get_object_or_404(
        Factura,
        id=factura_id,  
        cliente=cliente
    )

    # Igual que en la vista del admin
    detalles = DetalleFactura.objects.filter(factura=factura)

    return render(request, "clientes/cliente_factura_detalle.html", {
        "cliente": cliente,
        "factura": factura,
        "detalles": detalles,
    })


@cliente_required
def mis_datos(request):
    cliente = request.user.cliente_account.cliente

    if request.method == "POST":
        form = ClienteDatosForm(request.POST, instance=cliente)

        if form.is_valid():
            cliente_actualizado = form.save(commit=False)

            # Auditoría si la usas en tu modelo
            cliente_actualizado._auditoria_user = request.user

            cliente_actualizado.save()

            messages.success(request, "Datos actualizados correctamente.")
            return redirect("mis_datos")
        else:
            messages.error(request, "Por favor corrige los errores del formulario.")

    else:
        form = ClienteDatosForm(instance=cliente)

    return render(request, "clientes/mis_datos.html", {
        "cliente": cliente,
        "form": form,
    })

@cliente_required
def factura_descargar_cliente(request, factura_id):
    """
    Permite al cliente descargar su factura en PDF.
    Solo facturas ACTIVA son descargables.
    """

    # Obtener el Cliente a través del ClienteAccount del usuario
    try:
        cliente_account = request.user.cliente_account
        cliente = cliente_account.cliente
    except ClienteAccount.DoesNotExist:
        return HttpResponse("No tienes un cliente asociado.", status=403)

    # Obtener la factura del cliente
    factura = get_object_or_404(Factura, id=factura_id, cliente=cliente)

    # Bloquear si la factura no está activa
    if factura.estado != "ACTIVA":
        return render(
            request,
            "facturacion/factura_bloqueada.html",
            {"factura": factura},
            status=403
        )

    # Logo con ruta absoluta (WeasyPrint lo necesita)
    logo_url = request.build_absolute_uri(static("img/logo.png"))

    # Datos de la empresa
    empresa = {
        "nombre": "NovaBilling",
        "ruc": "8907826539023",
        "direccion": "Quito, Av Reinoso Rueda y Calle 8",
        "telefono": "0993395049",
        "correo": "contacto@novabilling.com",
    }

    # Detalles de la factura
    detalles = factura.detalles.all()

    # Renderizar la plantilla HTML de factura para PDF
    html_string = render_to_string(
        "facturacion/factura_pdf.html",
        {
            "factura": factura,
            "detalles": detalles,
            "empresa": empresa,
            "logo_url": logo_url,
        }
    )

    # Generar PDF con WeasyPrint
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()

    # Devolver PDF como descarga
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Factura_{factura.numero}.pdf"'

    return response


@cliente_required
def factura_imprimir_cliente(request, factura_id):

    try:

        factura = get_object_or_404(Factura, id=factura_id)

        # Obtener cliente desde la cuenta
        cliente = request.user.cliente_account.cliente

        # 🔐 Validar que la factura pertenece al cliente
        if factura.cliente != cliente:
            return render(
                request,
                "clientes/error_portal.html",
                {
                    "titulo": "Acceso no permitido",
                    "mensaje": "No tienes permiso para ver esta factura."
                },
                status=403
            )

        # 🔐 Validar estado
        if factura.estado != "ACTIVA":
            return render(
                request,
                "clientes/error_portal.html",
                {
                    "titulo": "Factura no disponible",
                    "mensaje": f"La factura {factura.numero} no puede imprimirse porque está anulada."
                },
                status=403
            )

        # Obtener detalles
        detalles = factura.detalles.all()

        if not detalles.exists():
            return render(
                request,
                "clientes/error_portal.html",
                {
                    "titulo": "Factura incompleta",
                    "mensaje": "Esta factura no tiene detalles registrados."
                },
                status=400
            )

        # Logo
        logo_url = request.build_absolute_uri(static("img/logo.png"))

        # Empresa
        empresa = {
            "nombre": "NovaBilling",
            "ruc": "8907826539023",
            "direccion": "Quito, Av Reinoso Rueda y Calle 8",
            "telefono": "0993395049",
            "correo": "contacto@novabilling.com",
        }

        html_string = render_to_string(
            "facturacion/factura_pdf.html",
            {
                "factura": factura,
                "detalles": detalles,
                "empresa": empresa,
                "logo_url": logo_url,
            }
        )

        html = HTML(
            string=html_string,
            base_url=request.build_absolute_uri()
        )

        pdf = html.write_pdf()

        response = HttpResponse(pdf, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'inline; filename="factura_{factura.id}.pdf"'
        )

        return response

    except Exception as e:

        print("ERROR GENERANDO FACTURA:", e)

        return render(
            request,
            "clientes/error_portal.html",
            {
                "titulo": "Error al generar la factura",
                "mensaje": "Ocurrió un problema al intentar generar el PDF."
            },
            status=500
        )