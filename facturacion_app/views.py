from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.utils.timezone import now
from django.utils import timezone
from django.http import JsonResponse
from django.db.models import Q
from django.core.serializers import serialize
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.conf import settings
from django.templatetags.static import static
import os
import random
from django.shortcuts import get_object_or_404
from datetime import datetime
import json
from django.contrib import messages
from usuarios.decorators import group_required



from clientes.models import Cliente
from productos.models import Producto
from .models import Factura, DetalleFactura

# ============================
# GENERAR CLAVE ACCESO SRI
# ============================

@group_required(["Vendedor", "SuperAdmin"])
def generar_clave_acceso_sri(
    fecha_emision,
    ruc,
    establecimiento,
    punto_emision,
    secuencial,
    ambiente="1",          # 1 = PRUEBAS
    tipo_comprobante="01"  # 01 = FACTURA
):
    fecha = fecha_emision.strftime("%d%m%Y")
    serie = f"{establecimiento}{punto_emision}"
    secuencial = str(secuencial).zfill(9)
    codigo_numerico = str(random.randint(10000000, 99999999))
    tipo_emision = "1"

    clave = (
        fecha +
        tipo_comprobante +
        ruc +
        ambiente +
        serie +
        secuencial +
        codigo_numerico +
        tipo_emision
    )

    # ===== MÓDULO 11 =====
    factores = [2, 3, 4, 5, 6, 7]
    suma = 0
    factor = 0

    for digito in reversed(clave):
        suma += int(digito) * factores[factor]
        factor = (factor + 1) % len(factores)

    mod = 11 - (suma % 11)

    if mod == 11:
        verificador = "0"
    elif mod == 10:
        verificador = "1"
    else:
        verificador = str(mod)

    return clave + verificador

# ============================
# IMPRIMIR FACTURA
# ============================

@group_required(["Vendedor", "SuperAdmin"])
def imprimir_factura(request, id):
    factura = get_object_or_404(Factura, id=id)

    # Proteger impresión
    if factura.estado != "ACTIVA":
        return render(
            request,
            "facturacion/factura_bloqueada.html",
            {
                "factura": factura
            },
            status=403
        )

    # Logo con ruta absoluta (WeasyPrint lo necesita)
    logo_url = request.build_absolute_uri(static("img/logo.png"))

    # Empresa
    empresa = {
        "nombre": "Super",
        "ruc": "99999999999999",
        "direccion": "Quito, Av Reinoso Rueda y Calle 8",
        "telefono": "0993395049",
        "correo": "contacto@empresa.com",
    }

    # Detalles
    detalles = factura.detalles.all()

    html_string = render_to_string(
        "facturacion/factura_pdf.html",
        {
            "factura": factura,
            "detalles": detalles,
            "empresa": empresa,
            "logo_url": logo_url,
        }
    )

    html = HTML(
        string=html_string,
        base_url=request.build_absolute_uri()
    )

    pdf = html.write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="factura_{factura.id}.pdf"'
    )

    return response



# ============================
# LISTAR FACTURAS
# ============================
@group_required(["Vendedor", "SuperAdmin"])
def lista_facturas(request):
    facturas = Factura.objects.all().order_by('-fecha')
    return render(request, 'facturacion/lista.html', {'facturas': facturas})


# ============================
# CREAR FACTURA
# ============================
@group_required(["Vendedor", "SuperAdmin"])
@transaction.atomic
def crear_factura(request):
    clientes = Cliente.objects.all()
    productos = Producto.objects.filter(estado=True)

    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        forma_pago = request.POST.get('forma_pago')

        if not cliente_id:
            messages.error(request, " Debes seleccionar un cliente.")
            return render(request, 'facturacion/crear.html', {'clientes': clientes, 'productos': productos, 'form_data': request.POST})

        if not forma_pago:
            messages.error(request, " Debes seleccionar una forma de pago.")
            return render(request, 'facturacion/crear.html', {'clientes': clientes, 'productos': productos, 'form_data': request.POST})

        productos_ids = request.POST.getlist('producto_id')
        cantidades = request.POST.getlist('cantidad')
        precios = request.POST.getlist('precio_unitario')
        descuentos = request.POST.getlist('descuento')

        if not productos_ids:
            messages.error(request, " Debes agregar al menos un producto a la factura.")
            return render(request, 'facturacion/crear.html', {'clientes': clientes, 'productos': productos, 'form_data': request.POST})

        if not (len(productos_ids) == len(cantidades) == len(precios) == len(descuentos)):
            messages.error(request, " Los detalles de la factura están incompletos.")
            return render(request, 'facturacion/crear.html', {'clientes': clientes, 'productos': productos, 'form_data': request.POST})

        # Número de factura simple
        ultimo = Factura.objects.count() + 1
        numero_factura = f"001-001-{ultimo:09d}"

        # Creamos factura “temporal” (si algo falla, la eliminamos)
        factura = Factura.objects.create(
            cliente_id=cliente_id,
            forma_pago=forma_pago,
            numero=numero_factura,
            fecha=now()
        )

        def fail(msg: str):
            # elimina lo creado en esta transacción y vuelve al form
            factura.delete()
            messages.error(request, msg)
            return render(request, 'facturacion/crear.html', {
                'clientes': clientes,
                'productos': productos,
                'form_data': request.POST
            })

        subtotal = 0
        subtotal_iva = 0
        subtotal_cero = 0
        descuento_total = 0

        for i in range(len(productos_ids)):
            # Parse seguro
            try:
                prod_id = int(productos_ids[i])
                cantidad = int((cantidades[i] or "0").strip())
                precio = float((precios[i] or "0").strip())
                desc = float((descuentos[i] or "0").strip())
            except (ValueError, TypeError):
                return fail(" Hay valores inválidos en los detalles de la factura.")

            # Validaciones por línea
            if cantidad <= 0:
                return fail(" La cantidad debe ser mayor a 0.")
            if precio < 0:
                return fail(" El precio no puede ser negativo.")
            if desc < 0:
                return fail(" El descuento no puede ser negativo.")

            # Producto + stock (bloquea fila para evitar carreras)
            prod = Producto.objects.select_for_update().get(id=prod_id)

            if not prod.estado:
                return fail(f" El producto '{prod.nombre}' no está disponible.")
            if prod.stock < cantidad:
                return fail(f" Stock insuficiente para '{prod.nombre}'. Disponible: {prod.stock}.")

            sub = (precio * cantidad) - desc
            if sub < 0:
                return fail(f" El subtotal de '{prod.nombre}' no puede ser negativo.")

            # Guardar detalle
            DetalleFactura.objects.create(
                factura=factura,
                producto=prod,
                cantidad=cantidad,
                precio_unitario=precio,
                descuento=desc,
                subtotal=sub
            )

            # Actualizar stock (audit)
            prod.stock -= cantidad
            prod._current_user = request.user
            prod._ip_address = request.META.get('REMOTE_ADDR')
            prod.save()

            descuento_total += desc
            subtotal += sub
            if prod.iva:
                subtotal_iva += sub
            else:
                subtotal_cero += sub

        iva_total = subtotal_iva * 0.12
        total = subtotal + iva_total

        factura.subtotal = subtotal
        factura.subtotal_iva = subtotal_iva
        factura.subtotal_cero = subtotal_cero
        factura.descuento_total = descuento_total
        factura.iva_total = iva_total
        factura.total = total

        factura.clave_acceso = generar_clave_acceso_sri(
            fecha_emision=factura.fecha,
            ruc="99999999999999",
            establecimiento="001",
            punto_emision="001",
            secuencial=ultimo,
            ambiente="1"
        )

        factura._current_user = request.user
        factura._ip_address = request.META.get('REMOTE_ADDR')
        factura.save()

        messages.success(request, f" Factura {factura.numero} creada correctamente.")
        return redirect('factura_lista')

    return render(request, 'facturacion/crear.html', {
        'clientes': clientes,
        'productos': productos
    })


# ============================
# VER FACTURA
# ============================
@group_required(["Vendedor", "SuperAdmin"])
def ver_factura(request, id):
    factura = get_object_or_404(Factura, id=id)
    detalles = DetalleFactura.objects.filter(factura=factura)

    return render(request, 'facturacion/ver.html', {
        'factura': factura,
        'detalles': detalles
    })

# ============================
# ANULAR FACTURA
# ============================
@group_required(["Vendedor", "SuperAdmin"])
@transaction.atomic
def anular_factura(request, id):
    factura = get_object_or_404(Factura, id=id)

    if factura.estado == "ANULADA":
        messages.info(request, " La factura ya se encuentra anulada.")
        return redirect('factura_lista')

    for det in factura.detalles.select_related('producto').select_for_update():
        prod = det.producto
        prod.stock += det.cantidad
        prod._current_user = request.user
        prod._ip_address = request.META.get('REMOTE_ADDR')
        prod.save()

    factura.estado = "ANULADA"
    factura._current_user = request.user
    factura._ip_address = request.META.get('REMOTE_ADDR')
    factura.save()

    messages.success(request, f" Factura {factura.numero} anulada correctamente.")
    return redirect('factura_lista')

# ============================
# AJAX: Buscar Producto
# ============================
@group_required(["Vendedor", "SuperAdmin"])
def buscar_producto(request):
    query = request.GET.get('q', '').strip()

    productos = Producto.objects.filter(
        Q(nombre__icontains=query) | Q(codigo__icontains=query),
        estado=True
    )[:10]  # limitar resultados

    data = [
        {
            'id': p.id,
            'codigo': p.codigo,
            'nombre': p.nombre,
            'precio': float(p.precio),
            'iva': p.iva,
            'stock': p.stock
        }
        for p in productos
    ]

    return JsonResponse({'resultados': data})


# ============================
# AJAX: Información del Cliente
# ============================
@group_required(["Vendedor", "SuperAdmin"])
def cliente_info(request, id):
    try:
        c = Cliente.objects.get(id=id)
        data = {
            'id': c.id,
            'identificacion': c.identificacion,
            'nombre': c.nombre_razon_social,
            'telefono': c.telefono,
            'direccion': c.direccion,
            'correo': c.correo,
        }
        return JsonResponse(data)

    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)
    

@group_required(["Vendedor", "SuperAdmin"])
def export_facturas_pdf(request):
    facturas = (
        Factura.objects
        .select_related("cliente")
        .all()
        .order_by("-fecha", "-id")
    )

    html_string = render_to_string(
        "facturacion/export_facturas_pdf.html",
        {
            "facturas": facturas,
            "fecha": timezone.localtime().strftime("%d/%m/%Y %H:%M"),
        }
    )

    pdf = HTML(
        string=html_string,
        base_url=request.build_absolute_uri("/")
    ).write_pdf()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="facturas.pdf"'
    return response


@group_required(["Vendedor", "SuperAdmin"])
def export_facturas_excel(request):
    facturas = (
        Factura.objects
        .select_related("cliente")
        .all()
        .order_by("-fecha", "-id")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # ❌ Quitamos "Subtotal 0"
    headers = [
        "#", "Número", "Fecha", "Cliente", "Forma de pago",
        "Subtotal", "Desc. total", "Subtotal IVA",
        "IVA 12%", "Total", "Estado"
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for i, f in enumerate(facturas, start=1):
        ws.append([
            i,
            f.numero,
            timezone.localtime(f.fecha).strftime("%d/%m/%Y %H:%M") if f.fecha else "",
            f.cliente.nombre_razon_social,
            f.forma_pago,
            float(f.subtotal or 0),
            float(f.descuento_total or 0),
            float(f.subtotal_iva or 0),
            float(f.iva_total or 0),
            float(f.total or 0),
            f.estado,
        ])

    # 💰 formato moneda
    # Subtotal → Total (columnas 6 a 10)
    for row in range(2, ws.max_row + 1):
        for col in range(6, 11):
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    # 📏 ancho columnas
    for col in range(1, len(headers) + 1):
        col_letter = get_column_letter(col)
        max_len = max(
            len(str(cell.value)) if cell.value else 0
            for cell in ws[col_letter]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    filename = f"facturas_{timezone.localtime().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



